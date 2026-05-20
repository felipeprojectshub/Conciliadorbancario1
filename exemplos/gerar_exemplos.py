"""
Gera 3 planilhas de exemplo com layouts distintos para demonstração do sistema.

Executar: python exemplos/gerar_exemplos.py
"""
from __future__ import annotations
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill

OUT = Path(__file__).parent


def gerar_alfa():
    """Layout Alfa: simples, coluna única de valor, inclui SISPAG e linha SDO."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extrato"
    ws.append(["Data", "Historico", "Valor"])
    dados = [
        ("02/01/2024", "SDO ANTERIOR",            "15000.00"),
        ("05/01/2024", "PIX RECEBIDO CLIENTE A",  "3500.00"),
        ("08/01/2024", "SISPAG LOTE 001",         "-4200.00"),
        ("10/01/2024", "TED FORNECEDOR B",         "-800.00"),
        ("12/01/2024", "PIX RECEBIDO CLIENTE C",   "1200.00"),
        ("15/01/2024", "IOF OPERACAO 123",           "-45.00"),
        ("20/01/2024", "PIX CLIENTE D",            "2800.00"),
        ("25/01/2024", "SISPAG LOTE 002",         "-3100.00"),
        ("31/01/2024", "SDO DO DIA",              "14355.00"),
    ]
    for row in dados:
        ws.append(row)

    # Planilha financeira correspondente
    ws2 = wb.create_sheet("Financeiro")
    ws2.append(["Data Venc", "Descricao", "Valor", "Classificacao"])
    fin = [
        ("05/01/2024", "Recebimento Cliente A - NF 001",   "3500.00",  "RECEITA SERVICOS"),
        ("08/01/2024", "Pgto Fornecedor X - SISPAG",      "-1400.00",  "CUSTO MERCADORIA"),
        ("08/01/2024", "Pgto Fornecedor Y - SISPAG",      "-2800.00",  "DESPESA OPERACIONAL"),
        ("10/01/2024", "Pgto Fornecedor B - TED",          "-800.00",  "DESPESA OPERACIONAL"),
        ("12/01/2024", "Recebimento Cliente C - NF 002",   "1200.00",  "RECEITA SERVICOS"),
        ("15/01/2024", "IOF Operacao 123",                  "-45.00",  "DESPESA FINANCEIRA"),
        ("20/01/2024", "Recebimento Cliente D - NF 003",  "2800.00",   "RECEITA SERVICOS"),
        ("25/01/2024", "Pgto Salarios - SISPAG",         "-3100.00",  "FOLHA PAGAMENTO"),
    ]
    for row in fin:
        ws2.append(row)

    wb.save(OUT / "exemplo_alfa.xlsx")
    print("  exemplo_alfa.xlsx gerado.")


def gerar_beta():
    """Layout Beta: 3 linhas de junk antes do header, colunas Debito/Credito separadas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extrato Bancario"

    # 3 linhas de lixo antes do header real
    ws.append(["BANCO EXEMPLO S.A."])
    ws.append(["Conta: 12345-6 | Agencia: 0001"])
    ws.append(["Periodo: 01/02/2024 a 28/02/2024"])
    ws.append(["Data", "Lancamento", "Debito", "Credito", "Saldo"])

    dados = [
        ("01/02/2024", "SALDO ANTERIOR",          "",         "8000.00", "8000.00"),
        ("05/02/2024", "PIX PGTO NF 201",      "1500.00",          "",   "6500.00"),
        ("07/02/2024", "TED RECB CLIENTE E",         "",     "4000.00", "10500.00"),
        ("10/02/2024", "BOLETO FORNEC F",        "750.00",          "",   "9750.00"),
        ("14/02/2024", "PIX PGTO NF 202",        "300.00",          "",   "9450.00"),
        ("20/02/2024", "CREDITO JUROS",                "",      "12.50",  "9462.50"),
        ("28/02/2024", "SALDO FINAL",                  "",         "",         ""),
    ]
    for row in dados:
        ws.append(row)

    # Financeiro
    ws2 = wb.create_sheet("Contas Pagar Receber")
    ws2.append(["Data", "Documento", "Fornecedor Cliente", "Debito", "Credito", "Centro Custo"])
    fin = [
        ("05/02/2024", "NF-201", "Fornecedor Alpha",    "1500.00", "",       "PRODUCAO"),
        ("07/02/2024", "REC-05", "Cliente E",           "",        "4000.00", "COMERCIAL"),
        ("10/02/2024", "NF-155", "Fornecedor Beta",      "750.00", "",       "ADMINISTRATIVO"),
        ("14/02/2024", "NF-202", "Fornecedor Gama",      "300.00", "",       "PRODUCAO"),
        ("20/02/2024", "JUROS",  "Banco Exemplo",        "",        "12.50",  "FINANCEIRO"),
    ]
    for row in fin:
        ws2.append(row)

    wb.save(OUT / "exemplo_beta.xlsx")
    print("  exemplo_beta.xlsx gerado. (use skip_rows=3 para o extrato)")


def gerar_gama():
    """Layout Gama: extrato CSV + financeiro xlsx com histórico em múltiplas colunas."""
    import csv

    # Extrato como CSV
    csv_path = OUT / "exemplo_gama_extrato.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["dt_movimento", "descricao_1", "descricao_2", "valor_lancamento"])
        writer.writerows([
            ("2024-03-04", "DOC/TED",        "EMPRESA DELTA LTDA",    "-2200.00"),
            ("2024-03-06", "PIX",            "RECEB EPSILON S.A.",    "5000.00"),
            ("2024-03-06", "PIX",            "RECEB ZETA ME",         "800.00"),
            ("2024-03-11", "TARIFA",         "MANUTENCAO CONTA",      "-35.00"),
            ("2024-03-15", "SISPAG",         "LOTE PGTOS MARCO",      "-6500.00"),
            ("2024-03-20", "DOC/TED",        "RECEB CLIENTE ETA",     "3300.00"),
        ])
    print(f"  {csv_path.name} gerado.")

    # Financeiro xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lancamentos"
    ws.append(["data_competencia", "tipo", "historico", "complemento", "valor_r$", "plano_contas", "centro_resultado"])
    fin = [
        ("04/03/2024", "SAIDA",   "DOC TED",    "Pgto Delta Ltda NF-310",     "-2200.00", "FORNECEDORES",     "OPERACIONAL"),
        ("06/03/2024", "ENTRADA", "PIX",         "Receb Epsilon Fatura 02",     "5000.00", "CLIENTES",         "COMERCIAL"),
        ("06/03/2024", "ENTRADA", "PIX",         "Receb Zeta Parcela 1",         "800.00", "CLIENTES",         "COMERCIAL"),
        ("11/03/2024", "SAIDA",   "TARIFA",      "Manut conta marco 2024",       "-35.00", "DESPESAS BANCARIAS","FINANCEIRO"),
        ("15/03/2024", "SAIDA",   "SISPAG",      "Folha Pagamento Marco",      "-4000.00", "FOLHA",            "RH"),
        ("15/03/2024", "SAIDA",   "SISPAG",      "Adiantamento 13o Salario",   "-2500.00", "FOLHA",            "RH"),
        ("20/03/2024", "ENTRADA", "DOC TED",     "Receb Cliente Eta NF-88",     "3300.00", "CLIENTES",         "COMERCIAL"),
    ]
    for row in fin:
        ws.append(row)
    wb.save(OUT / "exemplo_gama_financeiro.xlsx")
    print("  exemplo_gama_financeiro.xlsx gerado.")
    print("  DICA: mapeie 'tipo' + 'historico' + 'complemento' como historico multicolunas.")


if __name__ == "__main__":
    print("Gerando planilhas de exemplo...")
    gerar_alfa()
    gerar_beta()
    gerar_gama()
    print("\nPronto! Arquivos salvos em:", OUT.resolve())
