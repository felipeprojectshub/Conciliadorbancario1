"""
Geração do relatório Excel com 6 abas.
"""
from __future__ import annotations
import datetime
import io
from decimal import Decimal
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from plan.planilha_contabil import aplicar_depara_contabil_indexed, build_depara_index

from .normalize import (
    STATUS_CONCILIADO, STATUS_CONCILIADO_MANUAL,
    STATUS_REVISAR, STATUS_REVISAR_COLISAO,
    STATUS_IGNORADO_SEM_PAR, STATUS_SEM_PAREAMENTO,
    STATUS_IGNORADO_USUARIO,
    STATUS_PARCIAL, STATUS_PENDENTE_PARCIAL,
)

# Cores
COR_CONCILIADO     = "C6EFCE"  # verde claro
COR_MANUAL         = "CCFFCC"  # verde mais claro
COR_SEM_PAR        = "FFEB9C"  # amarelo
COR_REVISAR        = "FFC7CE"  # vermelho claro
COR_IGNORADO       = "D9D9D9"  # cinza
COR_CABECALHO      = "2F75B6"  # azul
COR_PARCIAL        = "FFD966"  # amarelo alaranjado — parcialmente conciliado
COR_PENDENTE       = "F4B942"  # laranja — pendente de conciliação parcial

_DATE_FMT = "DD/MM/YYYY"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _header_font() -> Font:
    return Font(bold=True, color="FFFFFF")


def _auto_width(ws, min_width=10, max_width=60, max_rows=1000):
    for col in ws.iter_cols(max_row=min(ws.max_row, max_rows)):
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_width), max_width)


def _apply_date_format(ws) -> None:
    """
    Aplica number_format DD/MM/YYYY a cada célula de data na última linha gravada.
    Gravar datetime.date diretamente (sem strftime) garante tipo Date no Excel,
    o que habilita filtro hierárquico Ano > Mês > Dia e ordena corretamente.
    """
    for cell in ws[ws.max_row]:
        if isinstance(cell.value, (datetime.date, datetime.datetime)):
            cell.number_format = _DATE_FMT


def _as_date(value) -> object:
    """Retorna datetime.date se o valor for um objeto de data; caso contrário devolve o valor original."""
    if isinstance(value, datetime.datetime):
        return value.date()
    return value


def _to_float(v) -> object:
    """Converte Decimal para float; passa outros valores sem alteração."""
    return float(v) if isinstance(v, Decimal) else v


def _display_status(status: object) -> str:
    s = str(status or "").strip()
    labels = {
        STATUS_CONCILIADO: "Conciliado",
        STATUS_CONCILIADO_MANUAL: "Conciliado manual",
        STATUS_REVISAR: "A revisar",
        STATUS_REVISAR_COLISAO: "A revisar - colisao",
        STATUS_IGNORADO_SEM_PAR: "Sem par financeiro",
        STATUS_SEM_PAREAMENTO: "Sem pareamento",
        STATUS_IGNORADO_USUARIO: "Ignorado pelo usuario",
        STATUS_PARCIAL: "Parcialmente conciliado",
        STATUS_PENDENTE_PARCIAL: "Pendente parcial",
    }
    return labels.get(s, s)


def _display_metodo(metodo: object) -> str:
    m = str(metodo or "").strip()
    if not m:
        return ""
    if m == "manual":
        return "Manual"
    if m == "manual:ignorado":
        return "Ignorado manualmente"
    if m.startswith("fora_confronto_recebimento"):
        return "Fora do confronto - recebimento"
    if m.startswith("fora_confronto_pagamento"):
        return "Fora do confronto - pagamento"
    if m.startswith("pendente:"):
        return "Pendente de conciliacao parcial"
    if m.startswith("bloqueado:"):
        return "Bloqueado para revisao"
    if "ambiguo" in m:
        if m.startswith("N:1"):
            return "Ambiguo - varios bancos para um financeiro"
        if "parcial" in m:
            return "Ambiguo - parcial"
        return "Ambiguo - um banco para varios financeiros"
    if "grupo grande" in m or "grupo limitado" in m:
        return "Limite de combinacoes atingido"
    if "tempo esgotado" in m:
        return "Tempo de busca esgotado"
    if m.startswith("1:1 D") and m != "1:1 D":
        return "Exato com variacao de data"
    if m == "1:1 D":
        return "Exato no dia"
    if m.startswith("1:N Dvar soma"):
        return _append_soma("Um banco para varios financeiros com variacao de data", m)
    if m.startswith("1:N D parcial soma"):
        return _append_soma("Parcial um banco para varios financeiros", m)
    if m.startswith("1:N D soma"):
        return _append_soma("Um banco para varios financeiros", m)
    if m.startswith("N:1 revisao_unica soma"):
        return _append_soma("Varios bancos para um financeiro - revisao unica", m)
    if m.startswith("1:N D revisao_unica soma"):
        return _append_soma("Um banco para varios financeiros - revisao unica", m)
    if m.startswith("N:1 soma"):
        return _append_soma("Varios bancos para um financeiro", m)
    return m


def _append_soma(label: str, metodo: str) -> str:
    try:
        qtd = str(metodo).split("soma=", 1)[1].split()[0]
    except Exception:
        qtd = ""
    return f"{label} ({qtd} itens)" if qtd else label


def _display_tipo_conciliacao(metodo: object, status: object = "") -> str:
    m = str(metodo or "").strip()
    s = str(status or "").strip()
    if s == STATUS_REVISAR or s == STATUS_REVISAR_COLISAO or "ambiguo" in m:
        return "Revisao manual"
    if s == STATUS_SEM_PAREAMENTO:
        return "Nao conciliado"
    if s == STATUS_IGNORADO_SEM_PAR:
        return "Financeiro sem par"
    if s == STATUS_IGNORADO_USUARIO or m == "manual:ignorado":
        return "Ignorado"
    if s == STATUS_PARCIAL or "parcial soma" in m:
        return "Parcial"
    if s == STATUS_PENDENTE_PARCIAL or m.startswith("pendente:"):
        return "Pendente parcial"
    if m == "manual" or s == STATUS_CONCILIADO_MANUAL:
        return "Manual"
    if m.startswith("fora_confronto_"):
        return "Fora do confronto"
    if m.startswith("N:1"):
        return "N:1"
    if m.startswith("1:N"):
        return "1:N"
    if m.startswith("1:1"):
        return "1:1"
    if s == STATUS_CONCILIADO:
        return "Conciliado"
    return ""


def _display_detail_header(col: str) -> str:
    labels = {
        "_id": "ID",
        "_data": "Data",
        "_valor": "Valor",
        "_historico": "Historico",
        "_classif": "Classificacao",
        "_status": "Situacao",
        "_metodo": "Criterio",
        "_ids_fin": "IDs Financeiro",
        "_id_bnk": "IDs Banco",
    }
    return labels.get(col, col)


def build_report(
    df_bnk: pd.DataFrame,
    df_fin: pd.DataFrame,
    depara: Optional[dict] = None,
    conta_banco: str = "",
    hist_mode: str = "Banco + Financeiro",
) -> bytes:
    """
    Constrói workbook Excel com 6 abas e retorna bytes.
    depara: dict {classif -> conta_contabil}
    hist_mode: "Banco + Financeiro" | "Somente bancário" | "Somente financeiro"
    """
    for col in ["_id", "_data", "_valor", "_historico", "_classif", "_status", "_metodo", "_id_bnk"]:
        if col not in df_fin.columns:
            df_fin[col] = ""

    for col in ["_id", "_data", "_valor", "_historico", "_classif", "_status", "_metodo", "_ids_fin"]:
        if col not in df_bnk.columns:
            df_bnk[col] = ""

    depara = depara or {}
    depara_index = build_depara_index(depara)
    wb = Workbook()
    wb.remove(wb.active)

    _build_alterdata(wb, df_bnk, df_fin, depara_index, conta_banco, hist_mode)
    _build_consolidado(wb, df_bnk, df_fin, depara_index, conta_banco)
    _build_extrato(wb, df_bnk)
    _build_financeiro(wb, df_fin)
    _build_sem_par_bnk(wb, df_bnk)
    _build_sem_par_fin(wb, df_fin)
    _build_resumo(wb, df_bnk, df_fin)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _resolve_classif(ids_fin_str: str, fin_by_id: dict) -> str:
    if not ids_fin_str:
        return ""
    ids = [x.strip() for x in str(ids_fin_str).split(";") if x.strip()]
    classifs = []
    for id_f in ids:
        rec = fin_by_id.get(id_f)
        if rec:
            c = str(rec.get("_classif", "")).strip()
            if c and c not in classifs:
                classifs.append(c)
    if len(classifs) == 0:
        return ""
    if len(classifs) == 1:
        return classifs[0]
    return "[MULTIPLAS: " + ", ".join(classifs) + "]"


def _resolve_historico_fin(ids_fin_str: str, fin_by_id: dict, sep: str = " - ") -> str:
    if not ids_fin_str:
        return ""
    ids = [x.strip() for x in str(ids_fin_str).split(";") if x.strip()]
    hists = []
    for id_f in ids:
        rec = fin_by_id.get(id_f)
        if rec:
            h = str(rec.get("_historico", "")).strip()
            if h and h not in hists:
                hists.append(h)
    return sep.join(hists)


def _concat_hist(hist_banco: str, hist_fin: str, sep: str = " - ") -> str:
    parts = [h for h in (hist_banco.strip(), hist_fin.strip()) if h]
    return sep.join(parts)


def _pick_hist(hist_banco: str, hist_fin: str, hist_mode: str) -> str:
    """Retorna o histórico conforme a preferência do usuário."""
    if hist_mode == "Somente bancário":
        return hist_banco
    if hist_mode == "Somente financeiro":
        return hist_fin or hist_banco  # fallback para bancário quando não há financeiro vinculado
    return _concat_hist(hist_banco, hist_fin)  # "Banco + Financeiro" (padrão)


def _build_alterdata(wb, df_bnk, df_fin, depara_index, conta_banco, hist_mode="Banco + Financeiro"):
    ws = wb.create_sheet("Importação Alterdata")
    headers = ["Histórico", "Nota Fiscal", "Data", "Nat. Cod", "Nat. Desc.", "Valor", "Débito", "Crédito"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _fill(COR_CABECALHO)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center")

    fin_by_id: dict = {
        rec["_id"]: rec
        for rec in df_fin[["_id", "_valor", "_historico", "_classif"]].to_dict("records")
    }

    for row in df_bnk.to_dict("records"):
        metodo = str(row.get("_metodo", ""))
        ids_fin_str = str(row.get("_ids_fin", ""))
        data_val = _as_date(row.get("_data", ""))
        hist_banco = str(row.get("_historico", "")).strip()

        is_expandable = (
            (("1:N D soma" in metodo or "1:N Dvar soma" in metodo or "parcial soma" in metodo)
             and "ambiguo" not in metodo)
            or metodo == "manual"
        ) and bool(ids_fin_str.strip())

        if is_expandable:
            ids_fin = [x.strip() for x in ids_fin_str.split(";") if x.strip()]
            bnk_valor = _to_float(row.get("_valor", ""))
            valid_fins = [(id_f, fin_by_id[id_f]) for id_f in ids_fin if id_f in fin_by_id]
            acumulado = 0.0
            for idx, (id_f, fin_row) in enumerate(valid_fins):
                is_last = (idx == len(valid_fins) - 1)
                classif = str(fin_row.get("_classif", "")).strip()
                if is_last:
                    valor_val = round(bnk_valor - acumulado, 2)
                else:
                    valor_val = _to_float(fin_row.get("_valor", ""))
                    acumulado += valor_val
                debito, credito, _ = aplicar_depara_contabil_indexed(
                    classif, valor_val, depara_index, conta_banco,
                )
                hist_fin = str(fin_row.get("_historico", "")).strip()
                ws.append([
                    _pick_hist(hist_banco, hist_fin, hist_mode),
                    None,
                    data_val,
                    None,
                    None,
                    valor_val,
                    debito,
                    credito,
                ])
                _apply_date_format(ws)
        else:
            classif = _resolve_classif(ids_fin_str, fin_by_id)
            valor_val = _to_float(row.get("_valor", ""))
            debito, credito, _ = aplicar_depara_contabil_indexed(
                classif, valor_val, depara_index, conta_banco,
            )
            hist_fin = _resolve_historico_fin(ids_fin_str, fin_by_id)
            ws.append([
                _pick_hist(hist_banco, hist_fin, hist_mode),
                None,
                data_val,
                None,
                None,
                valor_val,
                debito,
                credito,
            ])
            _apply_date_format(ws)

    _auto_width(ws)


def _build_consolidado(wb, df_bnk, df_fin, depara_index, conta_banco):
    ws = wb.create_sheet("Relatorio Consolidado")
    headers = [
        "Data", "Historico", "Valor", "Classificacao Financeira",
        "Historico Financeiro",
        "Tipo",
        "ID Banco", "ID Financeiro", "Criterio", "Situacao",
        "Debito", "Credito", "Status De x Para",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _fill(COR_CABECALHO)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center")

    status_cor = {
        STATUS_CONCILIADO:        COR_CONCILIADO,
        STATUS_CONCILIADO_MANUAL: COR_MANUAL,
        STATUS_SEM_PAREAMENTO:    COR_SEM_PAR,
        STATUS_REVISAR:           COR_REVISAR,
        STATUS_REVISAR_COLISAO:   COR_REVISAR,
        STATUS_IGNORADO_SEM_PAR:  COR_IGNORADO,
        STATUS_IGNORADO_USUARIO:  COR_IGNORADO,
        STATUS_PARCIAL:           COR_PARCIAL,
        STATUS_PENDENTE_PARCIAL:  COR_PENDENTE,
    }

    # Pre-indexa financeiros por _id — elimina scans lineares O(n) por linha
    fin_by_id: dict = {
        rec["_id"]: rec
        for rec in df_fin[["_id", "_valor", "_historico", "_classif"]].to_dict("records")
    }

    for row in df_bnk.to_dict("records"):
        metodo = str(row.get("_metodo", ""))
        ids_fin_str = str(row.get("_ids_fin", ""))
        status = str(row.get("_status", ""))
        cor = status_cor.get(status, "FFFFFF")

        data_val = _as_date(row.get("_data", ""))

        is_expandable = (
            (("1:N D soma" in metodo or "1:N Dvar soma" in metodo or "parcial soma" in metodo)
             and "ambiguo" not in metodo)
            or metodo == "manual"
        ) and bool(ids_fin_str.strip())

        if is_expandable:
            if "parcial soma" in metodo:
                tipo_expand = "Parcial 1:N"
            elif "Dvar" in metodo:
                tipo_expand = "Desmembrado 1:N ±D"
            elif "1:N" in metodo:
                tipo_expand = "Desmembrado 1:N"
            else:
                tipo_expand = "Manual"
            ids_fin = [x.strip() for x in ids_fin_str.split(";") if x.strip()]
            bnk_valor = _to_float(row.get("_valor", ""))
            valid_fins = [(id_f, fin_by_id[id_f]) for id_f in ids_fin if id_f in fin_by_id]
            acumulado = 0.0
            for idx, (id_f, fin_row) in enumerate(valid_fins):
                is_last = (idx == len(valid_fins) - 1)
                classif = str(fin_row.get("_classif", "")).strip()
                if is_last:
                    valor_val = round(bnk_valor - acumulado, 2)
                else:
                    valor_val = _to_float(fin_row.get("_valor", ""))
                    acumulado += valor_val
                debito, credito, status_depara = aplicar_depara_contabil_indexed(
                    classif,
                    valor_val,
                    depara_index,
                    conta_banco,
                )
                hist_fin = str(fin_row.get("_historico", "")).strip()

                linha = [
                    data_val,
                    str(row.get("_historico", "")),
                    valor_val,
                    classif,
                    hist_fin,
                    _display_tipo_conciliacao(metodo, status),
                    str(row.get("_id", "")),
                    id_f,
                    _display_metodo(metodo),
                    _display_status(status),
                    debito,
                    credito,
                    status_depara,
                ]
                ws.append(linha)
                _apply_date_format(ws)
                for cell in ws[ws.max_row]:
                    cell.fill = _fill(cor)
        else:
            classif = _resolve_classif(ids_fin_str, fin_by_id)
            tipo = ""
            if metodo == "1:1 D":
                tipo = "Exato 1:1"
            elif metodo.startswith("1:1 D"):
                tipo = "Exato 1:1 ±D"
            elif "parcial soma" in metodo:
                tipo = "Parcial 1:N"
            elif "1:N Dvar soma" in metodo:
                tipo = "Desmembrado 1:N ±D"
            elif "1:N D soma" in metodo:
                tipo = "Desmembrado 1:N"
            elif metodo.startswith("pendente:"):
                tipo = "Pendente Parcial"
            elif metodo == "manual":
                tipo = "Manual"

            hist_fin = _resolve_historico_fin(ids_fin_str, fin_by_id)
            valor_val = _to_float(row.get("_valor", ""))
            debito, credito, status_depara = aplicar_depara_contabil_indexed(
                classif,
                valor_val,
                depara_index,
                conta_banco,
            )

            linha = [
                data_val,
                str(row.get("_historico", "")),
                valor_val,
                classif,
                hist_fin,
                _display_tipo_conciliacao(metodo, status),
                str(row.get("_id", "")),
                ids_fin_str,
                _display_metodo(metodo),
                _display_status(status),
                debito,
                credito,
                status_depara,
            ]
            ws.append(linha)
            _apply_date_format(ws)
            for cell in ws[ws.max_row]:
                cell.fill = _fill(cor)

    _auto_width(ws)


def _build_extrato(wb, df_bnk):
    ws = wb.create_sheet("Extrato Bancario")
    cols_base = ["_id", "_data", "_valor", "_historico", "_status", "_metodo", "_ids_fin"]
    extras = [c for c in df_bnk.columns if not c.startswith("_")]
    headers = cols_base + extras
    display_headers = [_display_detail_header(c) for c in headers]
    ws.append(display_headers)
    for cell in ws[1]:
        cell.fill = _fill(COR_CABECALHO)
        cell.font = _header_font()
    for row in df_bnk.to_dict("records"):
        linha = []
        for c in headers:
            v = row.get(c, "")
            if isinstance(v, Decimal):
                v = float(v)
            elif isinstance(v, datetime.datetime):
                v = v.date()
            elif c == "_status":
                v = _display_status(v)
            elif c == "_metodo":
                v = _display_metodo(v)
            linha.append(v)
        ws.append(linha)
        _apply_date_format(ws)
    _auto_width(ws)


def _build_financeiro(wb, df_fin):
    ws = wb.create_sheet("Financeiro")
    cols_base = ["_id", "_data", "_valor", "_historico", "_classif", "_status", "_metodo", "_id_bnk"]
    extras = [c for c in df_fin.columns if not c.startswith("_")]
    headers = cols_base + extras
    display_headers = [_display_detail_header(c) for c in headers]
    ws.append(display_headers)
    for cell in ws[1]:
        cell.fill = _fill(COR_CABECALHO)
        cell.font = _header_font()
    for row in df_fin.to_dict("records"):
        linha = []
        for c in headers:
            v = row.get(c, "")
            if isinstance(v, Decimal):
                v = float(v)
            elif isinstance(v, datetime.datetime):
                v = v.date()
            elif c == "_status":
                v = _display_status(v)
            elif c == "_metodo":
                v = _display_metodo(v)
            linha.append(v)
        ws.append(linha)
        _apply_date_format(ws)
    _auto_width(ws)


def _build_sem_par_bnk(wb, df_bnk):
    ws = wb.create_sheet("Banco Sem Par")
    headers = ["ID Banco", "Data", "Valor", "Historico", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _fill(COR_CABECALHO)
        cell.font = _header_font()
    cor_por_status = {
        STATUS_SEM_PAREAMENTO:   COR_SEM_PAR,
        STATUS_REVISAR:          COR_REVISAR,
        STATUS_REVISAR_COLISAO:  COR_REVISAR,
        STATUS_PENDENTE_PARCIAL: COR_PENDENTE,
    }
    revisar_set = set(cor_por_status)
    for row in df_bnk.to_dict("records"):
        st = str(row.get("_status", ""))
        if st not in revisar_set:
            continue
        data_val = _as_date(row.get("_data", ""))
        valor_val = _to_float(row.get("_valor", ""))
        ws.append([
            str(row.get("_id", "")),
            data_val,
            valor_val,
            str(row.get("_historico", "")),
            st,
        ])
        _apply_date_format(ws)
        for cell in ws[ws.max_row]:
            cell.fill = _fill(cor_por_status.get(st, COR_SEM_PAR))
    _auto_width(ws)


def _build_sem_par_fin(wb, df_fin):
    ws = wb.create_sheet("Financeiro Sem Par")
    headers = ["ID Fin", "Data", "Valor", "Historico", "Classificacao", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _fill(COR_CABECALHO)
        cell.font = _header_font()
    revisar_set = {STATUS_IGNORADO_SEM_PAR, STATUS_REVISAR, STATUS_REVISAR_COLISAO}
    for row in df_fin.to_dict("records"):
        if str(row.get("_status", "")) not in revisar_set:
            continue
        data_val = _as_date(row.get("_data", ""))
        valor_val = _to_float(row.get("_valor", ""))
        ws.append([
            str(row.get("_id", "")),
            data_val,
            valor_val,
            str(row.get("_historico", "")),
            str(row.get("_classif", "")),
            str(row.get("_status", "")),
        ])
        _apply_date_format(ws)
        for cell in ws[ws.max_row]:
            cell.fill = _fill(COR_SEM_PAR)
    _auto_width(ws)


def _build_resumo(wb, df_bnk, df_fin):
    ws = wb.create_sheet("Resumo")
    ws.append(["Metrica", "Quantidade", "Valor Total"])
    for cell in ws[1]:
        cell.fill = _fill(COR_CABECALHO)
        cell.font = _header_font()

    # Converte coluna _valor para float uma única vez (evita loop com Decimal por linha).
    # .astype(float) garante dtype numérico mesmo em DataFrames vazios (apply preserva dtype str).
    bnk_vals = df_bnk["_valor"].apply(lambda v: float(v) if v else 0.0).astype(float)
    fin_vals = df_fin["_valor"].apply(lambda v: float(v) if v else 0.0).astype(float)

    def soma(status_series, vals_series, status_list):
        mask = status_series.isin(status_list)
        return int(mask.sum()), round(float(vals_series[mask].sum()), 2)

    metricas = [
        ("Conciliados (auto)",        df_bnk["_status"], bnk_vals, [STATUS_CONCILIADO]),
        ("Conciliados (manual)",      df_bnk["_status"], bnk_vals, [STATUS_CONCILIADO_MANUAL]),
        ("Parcialmente Conciliados",  df_bnk["_status"], bnk_vals, [STATUS_PARCIAL]),
        ("Pendentes Parciais",        df_bnk["_status"], bnk_vals, [STATUS_PENDENTE_PARCIAL]),
        ("A Revisar",                 df_bnk["_status"], bnk_vals, [STATUS_REVISAR, STATUS_REVISAR_COLISAO]),
        ("Banco Não Conciliado",      df_bnk["_status"], bnk_vals, [STATUS_SEM_PAREAMENTO]),
        ("Financeiro Sem Par",        df_fin["_status"], fin_vals, [STATUS_IGNORADO_SEM_PAR]),
        ("Ignorados (usuario)",       df_bnk["_status"], bnk_vals, [STATUS_IGNORADO_USUARIO]),
    ]

    for label, status_series, vals_series, statuses in metricas:
        qtd, total = soma(status_series, vals_series, statuses)
        ws.append([label, qtd, total])

    _auto_width(ws)
